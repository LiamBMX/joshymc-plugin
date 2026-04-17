"""Convert Quest Tracker.xlsx into quests.yml for JoshyMC plugin."""
import re
import sys
import io
import os
import math

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = "C:/Users/liama/Downloads/Quest_Tracker.xlsx"
OUTPUT_PATH = "C:/Users/liama/IdeaProjects/joshymc/src/main/resources/quests.yml"
COPY_PATH = "C:/Users/liama/IdeaProjects/joshymc/run/plugins/Joshymc/quests.yml"

SHEET_TO_CATEGORY = {
    "Mining": "MINING",
    "Slaying": "COMBAT",
    "Exploring": "EXPLORATION",
    "Farming": "FARMING",
    "Fishing": "FISHING",
    "Time Played": "EXPLORATION",
}

DIFFICULTY_MAP = {
    "★★☆☆☆ Easy": "EASY",
    "★★★☆☆ Medium": "MEDIUM",
    "★★★★☆ Hard": "HARD",
    "★★★★★ Expert": "LEGENDARY",
    "💀 Extreme": "LEGENDARY",
}

# Reward ranges: (money_min, money_max, xp_min, xp_max)
REWARD_RANGES = {
    "EASY":      (5000, 15000, 50, 150),
    "MEDIUM":    (20000, 80000, 200, 500),
    "HARD":      (100000, 500000, 500, 2000),
    "LEGENDARY": (500000, 5000000, 2000, 10000),
}

# --- Objective parsing ---

def parse_number(text):
    """Extract the first number from text, removing commas."""
    m = re.search(r'[\d,]+', text)
    if m:
        return int(m.group().replace(',', ''))
    return 0

MINING_PATTERNS = [
    (r'Mine .+ Coal', 'BREAK_BLOCK', 'COAL_ORE'),
    (r'Mine .+ Copper Ore', 'BREAK_BLOCK', 'COPPER_ORE'),
    (r'Mine .+ Iron Ore', 'BREAK_BLOCK', 'IRON_ORE'),
    (r'Mine .+ Lapis Ore', 'BREAK_BLOCK', 'LAPIS_ORE'),
    (r'Mine .+ Gold Ore', 'BREAK_BLOCK', 'GOLD_ORE'),
    (r'Mine .+ Redstone Ore', 'BREAK_BLOCK', 'REDSTONE_ORE'),
    (r'Mine .+ Diamond', 'BREAK_BLOCK', 'DIAMOND_ORE'),
    (r'Mine .+ Ancient Debris', 'BREAK_BLOCK', 'ANCIENT_DEBRIS'),
    (r'Mine .+ Obsidian', 'BREAK_BLOCK', 'OBSIDIAN'),
    (r'Mine .+ Stone', 'BREAK_BLOCK', 'STONE'),
    (r'Excavate .+ Blocks', 'BREAK_BLOCK', ''),
    (r'Collect .+ Gravel', 'BREAK_BLOCK', 'GRAVEL'),
    (r'Extract .+ Amethyst', 'BREAK_BLOCK', 'AMETHYST_CLUSTER'),
]

SLAYING_PATTERNS = [
    (r'Kill .+ Pigs?', 'KILL_MOB', 'PIG'),
    (r'Kill .+ Cows?', 'KILL_MOB', 'COW'),
    (r'Kill .+ Sheep', 'KILL_MOB', 'SHEEP'),
    (r'Kill .+ Chickens?', 'KILL_MOB', 'CHICKEN'),
    (r'Kill .+ Rabbits?', 'KILL_MOB', 'RABBIT'),
    (r'Kill .+ Skeletons?', 'KILL_MOB', 'SKELETON'),
    (r'Kill .+ Endermen', 'KILL_MOB', 'ENDERMAN'),
    (r'Kill .+ Blazes?', 'KILL_MOB', 'BLAZE'),
    (r'Kill .+ Husks?', 'KILL_MOB', 'HUSK'),
    (r'Defeat .+ Players?', 'KILL_PLAYER', ''),
    (r'Slay .+ Zombies?', 'KILL_MOB', 'ZOMBIE'),
    (r'Slay .+ Spiders?', 'KILL_MOB', 'SPIDER'),
    (r'Slay .+ Drowned', 'KILL_MOB', 'DROWNED'),
    (r'Defeat .+ Creepers?', 'KILL_MOB', 'CREEPER'),
    (r'Defeat .+ Witches?', 'KILL_MOB', 'WITCH'),
    (r'Defeat .+ Strays?', 'KILL_MOB', 'STRAY'),
    (r'Defeat .+ Elder Guardians?', 'KILL_MOB', 'ELDER_GUARDIAN'),
    (r'Defeat .+ Withers?$', 'KILL_MOB', 'WITHER'),  # $ to avoid matching "Wither Skeletons" if any
    (r'Defeat .+ Ender Dragons?', 'KILL_MOB', 'ENDER_DRAGON'),
    (r'Defeat .+ Wardens?', 'KILL_MOB', 'WARDEN'),
    (r'Defeat .+ Raid Captains?', 'KILL_MOB', 'PILLAGER'),
    (r'Defeat .*Boss', 'KILL_MOB', 'WARDEN'),
]

EXPLORING_PATTERNS = [
    (r'Walk .+ Blocks on Foot', 'WALK_DISTANCE', ''),
]

FARMING_PATTERNS = [
    (r'Harvest .+ Wheat', 'HARVEST_CROP', 'WHEAT'),
    (r'Grow .+ Carrots?', 'HARVEST_CROP', 'CARROTS'),
    (r'Plant .+ Seeds', 'PLACE_BLOCK', ''),
    (r'Grow .+ Melons?', 'HARVEST_CROP', 'MELON'),
    (r'Harvest .+ Pumpkins?', 'HARVEST_CROP', 'PUMPKIN'),
    (r'Collect .+ Apples?', 'BREAK_BLOCK', ''),
    (r'Brew .+ Potions?', 'CRAFT_ITEM', ''),
    (r'Collect .+ Mushrooms?', 'BREAK_BLOCK', 'BROWN_MUSHROOM'),
    (r'Harvest .+ Cocoa Beans?', 'HARVEST_CROP', 'COCOA'),
    (r'Collect .+ Kelp', 'HARVEST_CROP', 'KELP'),
    (r'Harvest .+ Sweet Berries?', 'HARVEST_CROP', 'SWEET_BERRY_BUSH'),
    (r'Harvest .+ Nether Wart', 'HARVEST_CROP', 'NETHER_WART'),
    (r'Harvest .+ Chorus Fruit', 'BREAK_BLOCK', 'CHORUS_PLANT'),
    (r'Harvest .+ Bamboo', 'BREAK_BLOCK', 'BAMBOO'),
    (r'Harvest .+ Sugar Cane', 'HARVEST_CROP', 'SUGAR_CANE'),
    (r'Collect .+ Chorus Fruit', 'BREAK_BLOCK', 'CHORUS_PLANT'),
]

FISHING_PATTERNS = [
    (r'Catch .+', 'CATCH_FISH', ''),
]

TIME_PLAYED_PATTERNS = [
    (r'Play for .+', 'TIME_PLAYED', ''),
]

CATEGORY_PATTERNS = {
    "MINING": MINING_PATTERNS,
    "COMBAT": SLAYING_PATTERNS,
    "EXPLORATION": EXPLORING_PATTERNS,
    "FARMING": FARMING_PATTERNS,
    "FISHING": FISHING_PATTERNS,
    "TIME_PLAYED": TIME_PLAYED_PATTERNS,
}

def parse_fishing_amount(objective):
    """Parse amount from fishing objective. For weight-based ('X lbs'), round down."""
    m = re.search(r'([\d,]+(?:\.\d+)?)', objective)
    if m:
        return int(float(m.group(1).replace(',', '')))
    return 0


def parse_time_played_hours(objective):
    """Parse 'Play for Xd Yh' into total seconds for TIME_PLAYED quests."""
    days = hours = 0
    m = re.search(r'(\d+)\s*d', objective)
    if m: days = int(m.group(1))
    m = re.search(r'(\d+)\s*h', objective)
    if m: hours = int(m.group(1))
    return days * 86400 + hours * 3600


def parse_objective(objective, category, sheet_name=""):
    """Parse objective text to (quest_type, target, amount)."""
    # Fishing: all quests are CATCH_FISH with empty target
    if sheet_name == "Fishing":
        amount = parse_fishing_amount(objective)
        return 'CATCH_FISH', '', amount

    # Time Played: map to TIME_PLAYED, value in seconds
    if sheet_name == "Time Played":
        amount = parse_time_played_hours(objective)
        return 'TIME_PLAYED', '', amount

    patterns = CATEGORY_PATTERNS[category]
    amount = parse_number(objective)

    # Boss objectives have a special format: "Defeat 2x Boss #3"
    boss_match = re.match(r'Defeat (?:(\d+)x )?Boss #(\d+)', objective)
    if boss_match:
        count = int(boss_match.group(1)) if boss_match.group(1) else 1
        return 'KILL_MOB', 'WARDEN', count

    for pattern, qtype, target in patterns:
        if re.match(pattern, objective):
            return qtype, target, amount

    print(f"  WARNING: No pattern matched for '{objective}' in {category}", file=sys.stderr)
    return 'BREAK_BLOCK', '', amount


def lerp(a, b, t):
    """Linear interpolation."""
    return int(a + (b - a) * t)


def compute_rewards(difficulty, quest_index_in_tier, total_in_tier):
    """Compute money and xp rewards scaled within the tier."""
    money_min, money_max, xp_min, xp_max = REWARD_RANGES[difficulty]
    if total_in_tier <= 1:
        t = 0.0
    else:
        t = quest_index_in_tier / (total_in_tier - 1)
    money = lerp(money_min, money_max, t)
    xp = lerp(xp_min, xp_max, t)
    return money, xp


def yaml_escape(s):
    """Escape a string for YAML if needed."""
    if any(c in s for c in ':#{}[]&*?|->!%@`"\','):
        return f'"{s}"'
    return f'"{s}"'


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # First pass: collect all quests grouped by difficulty tier per category
    all_quests = []  # list of dicts
    tier_counts = {}  # (category, difficulty) -> list of quest indices in all_quests

    for sheet_name, category in SHEET_TO_CATEGORY.items():
        ws = wb[sheet_name]
        # Determine quest ID prefix from sheet name
        if sheet_name == "Fishing":
            cat_prefix = "fishing"
        elif sheet_name == "Time Played":
            cat_prefix = "timeplayed"
        else:
            cat_prefix = category.lower()
            if cat_prefix == "combat":
                cat_prefix = "slaying"
            elif cat_prefix == "exploration":
                cat_prefix = "exploring"

        for row in range(3, ws.max_row + 1):
            quest_num = ws.cell(row, 1).value
            quest_name = ws.cell(row, 3).value  # Use objective as display name base
            objective_text = ws.cell(row, 3).value
            difficulty_raw = ws.cell(row, 4).value

            if quest_num is None or objective_text is None or difficulty_raw is None:
                continue

            quest_num = int(quest_num)
            difficulty = DIFFICULTY_MAP.get(difficulty_raw.strip(), "EASY")
            quest_type, target, amount = parse_objective(objective_text.strip(), category, sheet_name)

            quest_id = f"{cat_prefix}_{quest_num:03d}"
            quest = {
                "id": quest_id,
                "name": ws.cell(row, 2).value.strip(),
                "description": objective_text.strip(),
                "category": category,
                "type": quest_type,
                "target": target,
                "amount": amount,
                "difficulty": difficulty,
                "quest_num": quest_num,
            }

            idx = len(all_quests)
            all_quests.append(quest)

            key = (category, difficulty)
            if key not in tier_counts:
                tier_counts[key] = []
            tier_counts[key].append(idx)

    # Second pass: compute rewards
    for key, indices in tier_counts.items():
        category, difficulty = key
        total = len(indices)
        for i, idx in enumerate(indices):
            money, xp = compute_rewards(difficulty, i, total)
            all_quests[idx]["money"] = money
            all_quests[idx]["xp"] = xp

    # Write YAML
    lines = ["quests:"]
    for q in all_quests:
        lines.append(f"  {q['id']}:")
        lines.append(f"    name: {yaml_escape(q['name'])}")
        lines.append(f"    description: {yaml_escape(q['description'])}")
        lines.append(f"    category: {q['category']}")
        lines.append(f"    type: {q['type']}")
        lines.append(f"    target: {yaml_escape(q['target'])}")
        lines.append(f"    amount: {q['amount']}")
        lines.append(f"    difficulty: {q['difficulty']}")
        lines.append(f"    prerequisite: null")
        lines.append(f"    rewards:")
        lines.append(f"      money: {q['money']}")
        lines.append(f"      xp: {q['xp']}")
        lines.append(f"      items: []")

    content = "\n".join(lines) + "\n"

    # Write to primary location
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"Wrote {OUTPUT_PATH}")

    # Copy to runtime location
    os.makedirs(os.path.dirname(COPY_PATH), exist_ok=True)
    with open(COPY_PATH, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"Wrote {COPY_PATH}")

    # Summary
    cat_counts = {}
    for q in all_quests:
        cat_counts[q['category']] = cat_counts.get(q['category'], 0) + 1
    print(f"\nTotal quests: {len(all_quests)}")
    for cat, count in cat_counts.items():
        print(f"  {cat}: {count}")


if __name__ == "__main__":
    main()
